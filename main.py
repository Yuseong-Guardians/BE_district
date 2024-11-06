import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QPushButton, 
                             QFileDialog, QLabel, QListWidget, QListWidgetItem, 
                             QTabWidget, QTableWidget, QTableWidgetItem)
from PyQt5.QtWidgets import QHeaderView
from openpyxl.utils import get_column_letter
import pandas as pd
import os
from datetime import datetime


class ExcelMerger(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        
        self.setWindowTitle('엑셀 병합기')
        self.setGeometry(300, 300, 400, 300)
        self.setAcceptDrops(True)

        self.layout = QVBoxLayout()
        # 드래그 앤 드롭 안내 메시지
        self.label = QLabel('파일 목록')
        self.layout.addWidget(self.label)

        # 파일 리스트를 표시할 QListWidget
        self.file_list = QListWidget()
        self.layout.addWidget(self.file_list)

        # 파일 선택 버튼
        self.btn_select = QPushButton('엑셀 파일 선택', self)
        self.btn_select.clicked.connect(self.select_files)
        self.layout.addWidget(self.btn_select)

        # 병합 버튼
        self.btn_merge = QPushButton('병합 및 보기', self)
        self.btn_merge.clicked.connect(self.merge_files)
        self.layout.addWidget(self.btn_merge)

        # 추출 버튼
        self.btn_export = QPushButton('엑셀로 추출하기', self)
        self.btn_export.clicked.connect(self.export_to_excel)
        self.btn_export.setVisible(False)  # 초기에는 숨김
        self.layout.addWidget(self.btn_export)

        # 드래그 앤 드롭을 위한 설정
        self.setAcceptDrops(True)

        self.setLayout(self.layout)

        self.files = []
        self.merged_sheets = {}

    # 드래그된 파일이 들어왔을 때 호출되는 메서드
    def dragEnterEvent(self, event):
        print("Drag event detected")  # 이벤트 확인을 위한 출력문 추가
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            for url in urls:
                file_path = url.toLocalFile()
                print(f"Dragged file: {file_path}")  # 드래그된 파일 경로 출력
                if file_path.endswith(('.xlsx', '.xls')):
                    event.acceptProposedAction()
                else:
                    event.ignore()
        else:
            event.ignore()

    # 드롭된 파일을 처리하는 메서드
    def dropEvent(self, event):
        print("Drop event detected")  # 드롭 이벤트 확인용
        urls = event.mimeData().urls()
        for url in urls:
            file_path = url.toLocalFile()
            print(f"Dropped file: {file_path}")  # 드롭된 파일 경로 출력
            if file_path.endswith(('.xlsx', '.xls')):
                self.add_file(file_path)

    # 파일을 리스트에 추가하는 메서드
    def add_file(self, file_path):
        if file_path not in self.files:
            self.files.append(file_path)
            file_name = os.path.basename(file_path)
            item = QListWidgetItem(file_name)
            item.setToolTip(file_path)
            self.file_list.addItem(item)
        else:
            print(f"{file_path}는 이미 추가되었습니다.")

    # 파일 선택을 통해 엑셀 파일을 추가
    def select_files(self):
        selected_files, _ = QFileDialog.getOpenFileNames(self, '엑셀 파일 선택', '', '엑셀 파일 (*.xlsx *.xls)')
        for file in selected_files:
            self.add_file(file)

    def merge_files(self):
        if not self.files:
            print("파일을 먼저 선택하세요.")
            return

        # 병합할 시트 이름 목록
        sheet_names = ['현황', '신규자', '중지자']
        
        # 각각의 시트 데이터를 저장할 딕셔너리 초기화
        self.merged_sheets = {sheet_name: [] for sheet_name in sheet_names}

        # 각 파일의 시트를 읽어 병합
        for file in self.files:
            excel_data = pd.read_excel(file, sheet_name=None)  # 파일에서 모든 시트 읽기
            for sheet_name in sheet_names:
                if sheet_name in excel_data:
                    self.merged_sheets[sheet_name].append(excel_data[sheet_name])
                else:
                    print(f"파일 {file}에 {sheet_name} 시트가 없습니다.")

        # 시트별로 데이터를 병합
        for sheet_name in sheet_names:
            if self.merged_sheets[sheet_name]:
                self.merged_sheets[sheet_name] = pd.concat(self.merged_sheets[sheet_name], ignore_index=True)
            else:
                print(f"{sheet_name} 시트에 병합할 데이터가 없습니다.")

        # "현황" 시트에 "소급 필요" 열 추가 및 소급액 계산
        if '현황' in self.merged_sheets:
            status_data = self.merged_sheets['현황']
            status_data['소급 필요'] = ''  # 새로운 열 추가

            # 현재 날짜를 가져옵니다
            current_date = datetime.now()
            current_month_str = current_date.strftime("%Y-%m")
            
            def calculate_amount(row):
                try:
                    # 날짜 문자열 정리 및 변환
                    upload_date_str = row['등록일']
                    upload_date = datetime.strptime(upload_date_str, "%Y.%m.%d")
                    upload_month_str = upload_date.strftime("%Y-%m")

                    entry_date_str = row['전입일']
                    entry_date = datetime.strptime(entry_date_str, "%Y.%m.%d")

                    # 소급액 계산 및 소급 필요 정보 업데이트
                    if upload_month_str == current_month_str:
                        months_diff = (current_date.year - entry_date.year) * 12 + current_date.month - entry_date.month

                        # 사유가 전입이면 전입 당월은 이전 관할서에서 소급
                        if row['신규사유'] == '전입':
                            months_diff -= 1

                        # 소급 필요 정보 저장
                        retro_info = f"{months_diff}개월" if months_diff > 0 else ''
                        return retro_info 


                    else:
                        return ('')
                except ValueError as e:
                    print(f"Date conversion error: {e}")
                    return ('')

            # "소급 필요" 열 설정
            status_data['소급 필요'] = status_data.apply(lambda row: pd.Series(calculate_amount(row)), axis=1)

        # 병합 후 파일 선택 UI 제거
        self.clear_initial_ui()

        # 병합된 데이터를 UI로 보여줍니다
        self.show_merged_data()

        # 전체 화면으로 전환
        self.showMaximized()

        # 추출 버튼을 보이게 합니다
        self.btn_export.setVisible(True)

    # 초기 파일 선택 UI를 제거하는 메서드
    def clear_initial_ui(self):
        # 초기 UI 요소들을 숨김
        self.label.hide()
        self.file_list.hide()
        self.btn_select.hide()
        self.btn_merge.hide()

    # 병합된 데이터를 탭으로 보여주는 메서드
    def show_merged_data(self):
        self.tabs = QTabWidget()
        
        # 현황 탭 생성
        self.tab_status = QWidget()
        self.create_table_tab(self.tab_status, "현황")
        self.tabs.addTab(self.tab_status, "현황")
        
        # 신규자 탭 생성
        self.tab_new = QWidget()
        self.create_table_tab(self.tab_new, "신규자")
        self.tabs.addTab(self.tab_new, "신규자")

        # 중지자 탭 생성
        self.tab_stopped = QWidget()
        self.create_table_tab(self.tab_stopped, "중지자")
        self.tabs.addTab(self.tab_stopped, "중지자")

        self.layout.addWidget(self.tabs)

    # 각 탭에 병합된 데이터를 테이블로 표시하는 메서드
    def create_table_tab(self, tab, sheet_name):
        layout = QVBoxLayout()
        table = QTableWidget()
        self.load_data_to_table(sheet_name, table)
        layout.addWidget(table)
        tab.setLayout(layout)

        # 열 너비를 자동으로 조정
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    # 병합된 데이터를 QTableWidget에 로드하는 메서드
    def load_data_to_table(self, sheet_name, table_widget):
        data = self.merged_sheets.get(sheet_name)
        if data is not None:
            table_widget.setRowCount(data.shape[0])
            table_widget.setColumnCount(data.shape[1])
            table_widget.setHorizontalHeaderLabels(data.columns)

            for i in range(data.shape[0]):
                for j in range(data.shape[1]):
                    table_widget.setItem(i, j, QTableWidgetItem(str(data.iat[i, j])))


    # 엑셀로 추출하기 기능
    def export_to_excel(self):
        if not self.merged_sheets:
            print("먼저 파일을 병합하세요.")
            return

        # 현재 날짜를 가져옵니다.
        current_date_str = datetime.now().strftime("%Y%m")

        # 기본 파일 이름을 설정합니다.
        default_filename = f"참전유공자_{current_date_str}.xlsx"

        # 파일 저장 대화상자를 통해 파일 경로를 선택
        file_path, _ = QFileDialog.getSaveFileName(self, '엑셀 파일 저장', default_filename, '엑셀 파일 (*.xlsx)')
        if not file_path:
            return

        # 엑셀 파일로 저장
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, data in self.merged_sheets.items():
                # '계좌번호'와 '보훈번호' 열을 문자열로 변환
                if '계좌번호' in data.columns:
                    data['계좌번호'] = data['계좌번호'].astype(str)
                if '보훈번호' in data.columns:
                    data['보훈번호'] = data['보훈번호'].astype(str)

                # 엑셀 시트로 저장
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        # 열 크기 자동 조정을 위해 엑셀 파일을 다시 열어 처리
        from openpyxl import load_workbook

        workbook = load_workbook(file_path)

        for sheet_name in self.merged_sheets.keys():
            worksheet = workbook[sheet_name]

            # 각 열의 크기를 데이터에 맞춰 자동 조정
            for column_cells in worksheet.columns:
                max_length = 0
                column = get_column_letter(column_cells[0].column)  # A, B, C 등 열 번호를 가져옴
                for cell in column_cells:
                    try:  # 셀 값의 길이를 구해서 최대값을 찾음
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # 약간의 여유 공간과 보정을 추가
                worksheet.column_dimensions[column].width = adjusted_width

        # 저장 후 엑셀 파일 닫기
        workbook.save(file_path)
        workbook.close()

        print(f"파일이 저장되었습니다: {file_path}")
        try:
            os.startfile(file_path)  # 윈도우에서 파일을 기본 프로그램으로 열기
        except AttributeError:
            # 만약 os.startfile이 없는 플랫폼(예: Mac이나 Linux)에서는 다른 방법을 사용
            import subprocess
            subprocess.call(['open', file_path])  # Mac용
            # subprocess.call(['xdg-open', file_path])  # Linux용





if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelMerger()
    ex.show()
    sys.exit(app.exec_())
